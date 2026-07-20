#!/usr/bin/env python3
"""
Spreadsheet-to-Website Generator
Reads AI_Harness_Prompt_Kit_v39.xlsx and generates prompts.json and reference.json
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def extract_prompts_from_spreadsheet(xlsx_path: Path) -> list:
    """Extract prompt data from the spreadsheet."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    
    # Look for the main prompts sheet
    sheet_names = wb.sheetnames
    print(f"Available sheets: {sheet_names}")
    
    # Try to find the prompts sheet (common names)
    prompt_sheet = None
    for name in sheet_names:
        if 'prompt' in name.lower() or 'copy' in name.lower():
            prompt_sheet = wb[name]
            break
    
    if not prompt_sheet:
        # Try the first sheet
        prompt_sheet = wb[sheet_names[0]]
    
    print(f"Using sheet: {prompt_sheet.title}")
    
    # Read headers
    headers = []
    for cell in prompt_sheet[1]:
        if cell.value:
            headers.append(str(cell.value).strip())
    
    print(f"Headers found: {headers}")
    
    # Extract prompts
    prompts = []
    for row in prompt_sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        
        prompt_data = {}
        for i, header in enumerate(headers):
            if i < len(row) and row[i] is not None:
                prompt_data[header.lower().replace(' ', '_')] = str(row[i]).strip()
        
        if prompt_data.get('id') or prompt_data.get('prompt_id'):
            prompts.append(prompt_data)
    
    return prompts


def convert_to_prompts_json(prompts: list) -> list:
    """Convert extracted prompts to prompts.json format."""
    result = []
    
    for p in prompts:
        prompt_id = p.get('id') or p.get('prompt_id', '')
        if not prompt_id:
            continue
        
        # Map spreadsheet fields to JSON fields
        prompt_json = {
            "id": prompt_id,
            "seq": prompt_id.replace('P', ''),
            "name": p.get('name') or p.get('prompt_name', ''),
            "type": p.get('type') or p.get('prompt_type', 'BUILD'),
            "class": p.get('class') or p.get('prompt_class', ''),
            "sprintRole": p.get('sprint_role') or p.get('sprintrole', ''),
            "progress": p.get('progress', 'YES'),
            "useWhen": p.get('use_when') or p.get('usewhen', ''),
            "inspectFirst": p.get('inspect_first') or p.get('inspectfirst', ''),
            "expectedOutput": p.get('expected_output') or p.get('expectedoutput', ''),
            "nextStep": p.get('next_step') or p.get('nextstep', ''),
            "proofGate": p.get('proof_gate') or p.get('proofgate', ''),
            "color": p.get('color', 'Sky'),
            "copySheet": p.get('copy_sheet') or p.get('copysheet', ''),
            "category": p.get('category', 'standard'),
            "copyContent": p.get('copy_content') or p.get('copycontent', ''),
            "keywords": []
        }
        
        # Parse keywords if present
        keywords_str = p.get('keywords', '')
        if keywords_str:
            prompt_json["keywords"] = [k.strip() for k in keywords_str.split(',') if k.strip()]
        
        result.append(prompt_json)
    
    return result


def main():
    """Main entry point."""
    script_dir = Path(__file__).parent
    xlsx_path = script_dir / "AI_Harness_Prompt_Kit_v39.xlsx"
    
    if not xlsx_path.exists():
        print(f"Error: Spreadsheet not found at {xlsx_path}")
        sys.exit(1)
    
    print(f"Reading spreadsheet: {xlsx_path}")
    prompts = extract_prompts_from_spreadsheet(xlsx_path)
    print(f"Extracted {len(prompts)} prompts")
    
    # Convert to prompts.json format
    prompts_json = convert_to_prompts_json(prompts)
    
    # Write prompts.json
    output_path = script_dir / "prompts.json"
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(prompts_json, f, indent=2, ensure_ascii=False)
    
    print(f"Generated {output_path} with {len(prompts_json)} prompts")
    
    return prompts_json


if __name__ == "__main__":
    main()
