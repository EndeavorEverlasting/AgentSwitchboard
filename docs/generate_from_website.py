#!/usr/bin/env python3
"""
Website-to-Spreadsheet Generator
Reads prompts.json and reference.json and generates/upates the spreadsheet
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def load_json_data(script_dir: Path) -> tuple:
    """Load prompts.json and reference.json."""
    prompts_path = script_dir / "prompts.json"
    reference_path = script_dir / "reference.json"
    
    with open(prompts_path, 'r', encoding='utf-8') as f:
        prompts = json.load(f)
    
    with open(reference_path, 'r', encoding='utf-8') as f:
        reference = json.load(f)
    
    return prompts, reference


def create_workbook(prompts: list, reference: dict) -> openpyxl.Workbook:
    """Create Excel workbook from JSON data."""
    wb = openpyxl.Workbook()
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    gnhf_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Prompts
    ws_prompts = wb.active
    ws_prompts.title = "Prompts"
    
    # Headers
    headers = [
        "ID", "Seq", "Name", "Type", "Class", "Sprint Role", "Progress",
        "Use When", "Inspect First", "Expected Output", "Next Step",
        "Proof Gate", "Color", "Copy Sheet", "Category", "Copy Content", "Keywords"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws_prompts.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = border
    
    # Data rows
    for row_idx, prompt in enumerate(prompts, 2):
        row_data = [
            prompt.get('id', ''),
            prompt.get('seq', ''),
            prompt.get('name', ''),
            prompt.get('type', ''),
            prompt.get('class', ''),
            prompt.get('sprintRole', ''),
            prompt.get('progress', ''),
            prompt.get('useWhen', ''),
            prompt.get('inspectFirst', ''),
            prompt.get('expectedOutput', ''),
            prompt.get('nextStep', ''),
            prompt.get('proofGate', ''),
            prompt.get('color', ''),
            prompt.get('copySheet', ''),
            prompt.get('category', ''),
            prompt.get('copyContent', ''),
            ', '.join(prompt.get('keywords', []))
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws_prompts.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Highlight GNHF prompts
            if prompt.get('category') == 'gnhf':
                cell.fill = gnhf_fill
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws_prompts.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    # Sheet 2: Reference - GNHF Workflow
    ws_workflow = wb.create_sheet("GNHF Workflow")
    
    workflow_headers = [
        "Use Case", "Prompt", "Git Mode", "Iterations", "Token Cap",
        "Outcome", "Use After", "Stop Gate", "Paste Into"
    ]
    
    for col, header in enumerate(workflow_headers, 1):
        cell = ws_workflow.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = border
    
    for row_idx, item in enumerate(reference.get('gnhfWorkflow', []), 2):
        row_data = [
            item.get('useCase', ''),
            item.get('prompt', ''),
            item.get('gitMode', ''),
            item.get('iterations', ''),
            item.get('tokenCap', ''),
            item.get('outcome', ''),
            item.get('useAfter', ''),
            item.get('stopGate', ''),
            item.get('pasteInto', '')
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws_workflow.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Sheet 3: Reference - Variables
    ws_variables = wb.create_sheet("Variables")
    
    var_headers = ["Variable", "Meaning", "Example"]
    
    for col, header in enumerate(var_headers, 1):
        cell = ws_variables.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = border
    
    for row_idx, item in enumerate(reference.get('variables', []), 2):
        row_data = [
            item.get('variable', ''),
            item.get('meaning', ''),
            item.get('example', '')
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws_variables.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Sheet 4: Reference - Prompt Sequence
    ws_sequence = wb.create_sheet("Prompt Sequence")
    
    seq_headers = [
        "Seq", "Prompt ID", "Moment", "Use It For", "Do Not Use When",
        "Produces", "Gate", "Then", "Mutates Repo", "Authority",
        "Proof Ceiling", "Copy Safe Sheet"
    ]
    
    for col, header in enumerate(seq_headers, 1):
        cell = ws_sequence.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = border
    
    for row_idx, item in enumerate(reference.get('promptSequence', []), 2):
        row_data = [
            item.get('seq', ''),
            item.get('promptId', ''),
            item.get('moment', ''),
            item.get('useItFor', ''),
            item.get('doNotUseWhen', ''),
            item.get('produces', ''),
            item.get('gate', ''),
            item.get('then', ''),
            item.get('mutatesRepo', ''),
            item.get('authority', ''),
            item.get('proofCeiling', ''),
            item.get('copySafeSheet', '')
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws_sequence.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    return wb


def main():
    """Main entry point."""
    script_dir = Path(__file__).parent
    
    print("Loading JSON data...")
    prompts, reference = load_json_data(script_dir)
    print(f"Loaded {len(prompts)} prompts")
    
    print("Creating Excel workbook...")
    wb = create_workbook(prompts, reference)
    
    # Save workbook
    output_path = script_dir / "AI_Harness_Prompt_Kit_v39.xlsx"
    wb.save(output_path)
    
    print(f"Generated {output_path}")
    
    return output_path


if __name__ == "__main__":
    main()
